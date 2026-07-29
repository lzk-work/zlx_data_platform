"""Excel 输入读取。"""

from __future__ import annotations

from pathlib import Path
from typing import Callable, TypeVar

from openpyxl import load_workbook

from .models import (
    DailyFirstOrderRow,
    FirstCategoryCode,
    HistoricalOrderedPlatformSku,
    ProductSkuMaster,
    SkuPlatformMapping,
    UploadedProductSku,
    WorkbookRow,
)
from .normalizer import clean_source_url, clean_spec, normalize_sku, normalize_text

T = TypeVar("T")

DAILY_COLUMNS = ["平台SKU", "初始商品SKU", "订单号", "平台渠道", "店铺账号", "出单时间", "校正后货源链接", "校正后规格", "备注"]
PRODUCT_MASTER_COLUMNS = ["商品SKU", "货源链接", "规格"]
FIRST_CATEGORY_CODE_COLUMNS = ["code"]
UPLOADED_COLUMNS = ["商品SKU", "首次上传时间", "最后更新时间", "备注"]
HISTORICAL_COLUMNS = ["平台SKU", "订单号", "平台渠道", "店铺账号", "首次出单时间", "首次处理时间", "处理批次", "备注"]
MAPPING_COLUMNS = ["商品SKU", "平台SKU", "绑定时间", "最后更新时间", "绑定来源", "备注"]


def _read_rows(path: Path, required_columns: list[str], sheet_name: str | None = None) -> list[tuple[int, WorkbookRow]]:
    """读取 Excel 首个 sheet 或指定 sheet。"""
    if not path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {path}")
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Excel 文件没有表头: {path}")

    headers = [normalize_text(value) for value in rows[0]]
    missing = [column for column in required_columns if column not in headers]
    if missing:
        raise ValueError(f"{path} 缺少字段: {', '.join(missing)}")

    result: list[tuple[int, WorkbookRow]] = []
    for index, values in enumerate(rows[1:], start=2):
        row = {header: normalize_text(value) for header, value in zip(headers, values, strict=False) if header}
        if any(row.get(column, "") for column in required_columns):
            result.append((index, row))
    return result


def _load_table(path: Path, columns: list[str], factory: Callable[[int, WorkbookRow], T], sheet_name: str | None = None) -> list[T]:
    """读取表格并转换为模型。"""
    return [factory(row_number, row) for row_number, row in _read_rows(path, columns, sheet_name)]


def load_daily_input(path: Path, sheet_name: str | None = None) -> list[DailyFirstOrderRow]:
    return _load_table(path, DAILY_COLUMNS, _daily_row, sheet_name)


def load_product_sku_master(path: Path, sheet_name: str | None = None) -> list[ProductSkuMaster]:
    return _load_table(path, PRODUCT_MASTER_COLUMNS, _product_master_row, sheet_name)


def load_first_category_codes(path: Path, sheet_name: str | None = None) -> list[FirstCategoryCode]:
    return _load_table(path, FIRST_CATEGORY_CODE_COLUMNS, _first_category_code_row, sheet_name)


def load_uploaded_product_skus(path: Path, sheet_name: str | None = None) -> list[UploadedProductSku]:
    return _load_table(path, UPLOADED_COLUMNS, _uploaded_row, sheet_name)


def load_historical_ordered_platform_skus(path: Path, sheet_name: str | None = None) -> list[HistoricalOrderedPlatformSku]:
    return _load_table(path, HISTORICAL_COLUMNS, _historical_row, sheet_name)


def load_sku_platform_mapping(path: Path, sheet_name: str | None = None) -> list[SkuPlatformMapping]:
    return _load_table(path, MAPPING_COLUMNS, _mapping_row, sheet_name)


def _daily_row(row_number: int, row: WorkbookRow) -> DailyFirstOrderRow:
    return DailyFirstOrderRow(
        row_number,
        normalize_sku(row.get("平台SKU")),
        normalize_sku(row.get("初始商品SKU")),
        normalize_text(row.get("订单号")),
        normalize_text(row.get("平台渠道")),
        normalize_text(row.get("店铺账号")),
        normalize_text(row.get("出单时间")),
        clean_source_url(row.get("校正后货源链接")),
        clean_spec(row.get("校正后规格")),
        normalize_text(row.get("备注")),
        source_image_url=normalize_text(row.get("图片链接") or row.get("货源图片链接")),
        purchase_price=normalize_text(row.get("采购价/￥")),
        weight_g=normalize_text(row.get("重量/g")),
        length_cm=normalize_text(row.get("长/cm") or row.get("长")),
        width_cm=normalize_text(row.get("宽/cm") or row.get("宽")),
        height_cm=normalize_text(row.get("高/cm") or row.get("高")),
        color=normalize_text(row.get("颜色")),
        material=normalize_text(row.get("材质")),
        quantity=normalize_text(row.get("数量")),
        chinese_customs_name=normalize_text(row.get("中文报关名")),
        first_level_category=normalize_text(row.get("一级类目")),
        category_code=normalize_text(row.get("类目代号")),
        temp_sku=normalize_text(row.get("临时SKU")),
        supplier=normalize_text(row.get("供应商")),
    )


def _product_master_row(_row_number: int, row: WorkbookRow) -> ProductSkuMaster:
    return ProductSkuMaster(
        normalize_sku(row.get("商品SKU")),
        normalize_text(row.get("货源链接")),
        normalize_text(row.get("规格")),
        normalize_text(row.get("长/cm") or row.get("长")),
        normalize_text(row.get("宽/cm") or row.get("宽")),
        normalize_text(row.get("高/cm") or row.get("高")),
        source_image_url=normalize_text(row.get("货源图片链接") or row.get("图片链接")),
        purchase_price=normalize_text(row.get("采购价/￥")),
        weight_g=normalize_text(row.get("重量/g") or row.get("重量")),
        color=normalize_text(row.get("颜色")),
        material=normalize_text(row.get("材质")),
        quantity=normalize_text(row.get("数量")),
        chinese_customs_name=normalize_text(row.get("中文报关名")),
        first_level_category=normalize_text(row.get("一级类目")),
        category_code=normalize_text(row.get("类目代号")),
        temp_sku=normalize_text(row.get("临时SKU")),
        supplier=normalize_text(row.get("供应商")),
        note=normalize_text(row.get("note") or row.get("备注")),
    )


def _uploaded_row(_row_number: int, row: WorkbookRow) -> UploadedProductSku:
    return UploadedProductSku(normalize_sku(row.get("商品SKU")), normalize_text(row.get("首次上传时间")), normalize_text(row.get("最后更新时间")), normalize_text(row.get("备注")))


def _first_category_code_row(_row_number: int, row: WorkbookRow) -> FirstCategoryCode:
    return FirstCategoryCode(
        first_category=normalize_text(row.get("first_category") or row.get("一级类目")),
        first_category_chinese=normalize_text(row.get("first_category_chinese") or row.get("一级类目中文")),
        code=normalize_text(row.get("code") or row.get("类目代号")),
    )


def _historical_row(_row_number: int, row: WorkbookRow) -> HistoricalOrderedPlatformSku:
    return HistoricalOrderedPlatformSku(normalize_sku(row.get("平台SKU")), normalize_text(row.get("订单号")), normalize_text(row.get("平台渠道")), normalize_text(row.get("店铺账号")), normalize_text(row.get("首次出单时间")), normalize_text(row.get("首次处理时间")), normalize_text(row.get("处理批次")), normalize_text(row.get("备注")))


def _mapping_row(_row_number: int, row: WorkbookRow) -> SkuPlatformMapping:
    return SkuPlatformMapping(normalize_sku(row.get("商品SKU")), normalize_sku(row.get("平台SKU")), normalize_text(row.get("绑定时间")), normalize_text(row.get("最后更新时间")), normalize_text(row.get("绑定来源")), normalize_text(row.get("备注")))
