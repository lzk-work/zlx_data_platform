from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from apps.sku_mapping_poc.src.exporter import export_process_output
from apps.sku_mapping_poc.src.exporter import publish_state_files
from apps.sku_mapping_poc.src.loader import (
    load_daily_input,
    load_historical_ordered_platform_skus,
    load_product_sku_master,
    load_sku_platform_mapping,
    load_uploaded_product_skus,
)
from apps.sku_mapping_poc.src.models import AppSettings, UploadedProductSku
from apps.sku_mapping_poc.src.processor import process_sku_mapping


def settings(tmp_path: Path) -> AppSettings:
    return AppSettings(
        project_root=tmp_path,
        daily_input=tmp_path / "每日出单平台SKU输入表.xlsx",
        product_sku_master=tmp_path / "商品SKU总库.xlsx",
        uploaded_product_skus=tmp_path / "已上传商品SKU产品表.xlsx",
        historical_ordered_platform_skus=tmp_path / "历史出单平台SKU表.xlsx",
        product_sku_platform_sku_mapping=tmp_path / "商品SKU-平台SKU映射关系表.xlsx",
        output_dir=tmp_path / "output",
    )


def write_xlsx(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def values(path: Path) -> list[list[str | None]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def test_excel_inputs_export_expected_outputs(tmp_path: Path) -> None:
    config = settings(tmp_path)
    write_xlsx(
        config.daily_input,
        ["平台SKU", "初始商品SKU", "订单号", "平台渠道", "店铺账号", "出单时间", "校正后货源链接", "校正后规格", "类目代号", "备注"],
        [
            ["PSKU_UPD", "SKU000001", "O1", "Amazon-US", "shop-a", "2026-07-22 10:00:00", "https://detail.1688.com/offer/100001.html?spm=abc", "spec-1", "", "更新已上传"],
            ["PSKU_NEW", "SKU000002", "O2", "Amazon-US", "shop-a", "2026-07-22 11:00:00", "https://detail.1688.com/offer/100002.html?spm=abc", "spec-2----标题内容", "", "新增未上传"],
            ["PSKU_GEN", "SKU000001", "O3", "Amazon-US", "shop-a", "2026-07-22 12:00:00", "https://detail.1688.com/offer/199999.html?spm=abc", "spec-new----标题内容", "JT", "生成新SKU"],
            ["PSKU_HIS", "SKU000001", "O4", "Amazon-US", "shop-a", "2026-07-22 13:00:00", "ignored", "ignored", "", "历史跳过"],
        ],
    )
    write_xlsx(
        config.product_sku_master,
        ["商品SKU", "货源链接", "规格", "长", "宽", "高"],
        [["SKU000001", "https://detail.1688.com/offer/100001.html", "spec-1", "1", "2", "3"], ["SKU000002", "https://detail.1688.com/offer/100002.html", "spec-2", "4", "5", "6"]],
    )
    write_xlsx(config.uploaded_product_skus, ["商品SKU", "首次上传时间", "最后更新时间", "备注"], [["SKU000001", "", "", ""]])
    write_xlsx(
        config.historical_ordered_platform_skus,
        ["平台SKU", "订单号", "平台渠道", "店铺账号", "首次出单时间", "首次处理时间", "处理批次", "备注"],
        [["PSKU_HIS", "OH", "Amazon-US", "shop-a", "2026-07-21", "2026-07-21", "OLD", ""]],
    )
    write_xlsx(
        config.product_sku_platform_sku_mapping,
        ["商品SKU", "平台SKU", "绑定时间", "最后更新时间", "绑定来源", "备注"],
        [["SKU000001", "PSKU_OLD", "", "", "初始导入", ""]],
    )

    output = process_sku_mapping(
        settings=config,
        batch_id="20260708_000000",
        daily_rows=load_daily_input(config.daily_input),
        product_master_rows=load_product_sku_master(config.product_sku_master),
        uploaded_rows=load_uploaded_product_skus(config.uploaded_product_skus),
        historical_rows=load_historical_ordered_platform_skus(config.historical_ordered_platform_skus),
        mapping_rows=load_sku_platform_mapping(config.product_sku_platform_sku_mapping),
    )
    output_dir = export_process_output(config, output, "20260708_000000")

    assert output.summary is not None
    assert output.summary.first_order_processed == 3
    assert output.summary.historical_skipped == 1
    assert output.summary.generated_product_skus == 1

    erp_update = values(output_dir / "ERP更新表.xlsx")
    assert erp_update[0][:19] == [
        "*SKU\n(必填)",
        "平台SKU",
        "识别码",
        "中文名称",
        "英文名称",
        "分类ID",
        "图片URL\n（必须以http://或https：//开头）",
        "商品净重\n（g）",
        "采购参考价\n（RMB）",
        "采购员\n（输入子账号姓名或名称）",
        "长（cm）",
        "宽（cm）",
        "高（cm）",
        "来源URL\n（必须以http://或https：//开头）",
        "备注",
        "英文报关名",
        "中文报关名",
        "申报重量\n(g)",
        "申报金额\n（USD）",
    ]
    assert erp_update[1][0] == "SKU000001"
    assert erp_update[1][1] == "PSKU_OLD\nPSKU_UPD"

    erp_new = values(output_dir / "ERP新增表.xlsx")
    new_by_sku = {row[0]: row for row in erp_new[1:]}
    assert new_by_sku["SKU000002"][1] == "PSKU_NEW"
    assert new_by_sku["JT_260708_1"][1] == "PSKU_GEN"
    assert new_by_sku["JT_260708_1"][3] == "spec-new"
    assert new_by_sku["JT_260708_1"][13] == "https://detail.1688.com/offer/199999.html"

    latest_uploaded = values(output_dir / "最新已上传商品SKU产品表.xlsx")
    uploaded_skus = {row[0] for row in latest_uploaded[1:]}
    assert {"SKU000001", "SKU000002", "JT_260708_1"}.issubset(uploaded_skus)

    latest_history = values(output_dir / "最新历史出单平台SKU表.xlsx")
    history_skus = {row[0] for row in latest_history[1:]}
    assert {"PSKU_HIS", "PSKU_UPD", "PSKU_NEW", "PSKU_GEN"}.issubset(history_skus)

    logs = values(output_dir / "处理日志表.xlsx")
    logged_platform_skus = {row[2] for row in logs[1:]}
    assert logged_platform_skus == {"PSKU_UPD", "PSKU_NEW", "PSKU_GEN"}


def test_initial_run_can_start_without_state_tables(tmp_path: Path) -> None:
    config = settings(tmp_path)
    write_xlsx(
        config.daily_input,
        ["平台SKU", "初始商品SKU", "订单号", "平台渠道", "店铺账号", "出单时间", "校正后货源链接", "校正后规格", "备注"],
        [["PSKU_FIRST", "SKU000001", "O1", "Amazon-US", "shop-a", "2026-07-23 10:00:00", "https://detail.1688.com/offer/100001.html?spm=abc", "spec-1", "initial run"]],
    )
    write_xlsx(
        config.product_sku_master,
        ["商品SKU", "货源链接", "规格", "长", "宽", "高"],
        [["SKU000001", "https://detail.1688.com/offer/100001.html", "spec-1", "1", "2", "3"]],
    )

    output = process_sku_mapping(
        settings=config,
        batch_id="BATCH_INITIAL",
        daily_rows=load_daily_input(config.daily_input),
        product_master_rows=load_product_sku_master(config.product_sku_master),
        uploaded_rows=[],
        historical_rows=[],
        mapping_rows=[],
    )
    output_dir = export_process_output(config, output, "BATCH_INITIAL")

    assert output.summary is not None
    assert output.summary.first_order_processed == 1
    assert [row.product_sku for row in output.erp_new_rows] == ["SKU000001"]
    assert (output_dir / "最新已上传商品SKU产品表.xlsx").exists()
    assert (output_dir / "最新历史出单平台SKU表.xlsx").exists()
    assert (output_dir / "最新商品SKU-平台SKU映射关系表.xlsx").exists()


def test_publish_state_files_updates_next_run_inputs(tmp_path: Path) -> None:
    config = settings(tmp_path)
    output_dir = tmp_path / "output" / "BATCH_PUBLISH"
    output_dir.mkdir(parents=True)
    write_xlsx(output_dir / "最新已上传商品SKU产品表.xlsx", ["商品SKU", "首次上传时间", "最后更新时间", "备注"], [["SKU_NEW", "", "", ""]])
    write_xlsx(
        output_dir / "最新历史出单平台SKU表.xlsx",
        ["平台SKU", "订单号", "平台渠道", "店铺账号", "首次出单时间", "首次处理时间", "处理批次", "备注"],
        [["PSKU_NEW", "O1", "Amazon-US", "shop-a", "", "", "", ""]],
    )
    write_xlsx(
        output_dir / "最新商品SKU-平台SKU映射关系表.xlsx",
        ["商品SKU", "平台SKU", "绑定时间", "最后更新时间", "绑定来源", "备注"],
        [["SKU_NEW", "PSKU_NEW", "", "", "", ""]],
    )
    write_xlsx(config.uploaded_product_skus, ["商品SKU", "首次上传时间", "最后更新时间", "备注"], [["SKU_OLD", "", "", ""]])

    publish_state_files(config, output_dir, "BATCH_PUBLISH")

    assert values(config.uploaded_product_skus)[1][0] == "SKU_NEW"
    assert values(config.historical_ordered_platform_skus)[1][0] == "PSKU_NEW"
    assert values(config.product_sku_platform_sku_mapping)[1][:2] == ["SKU_NEW", "PSKU_NEW"]
    assert not list(tmp_path.glob("*.bak_BATCH_PUBLISH.xlsx"))


def test_published_state_makes_next_run_skip_same_platform_sku(tmp_path: Path) -> None:
    config = settings(tmp_path)
    write_xlsx(
        config.daily_input,
        ["平台SKU", "初始商品SKU", "订单号", "平台渠道", "店铺账号", "出单时间", "校正后货源链接", "校正后规格", "备注"],
        [["PSKU_FIRST", "SKU000001", "O1", "Amazon-US", "shop-a", "2026-07-23 10:00:00", "https://detail.1688.com/offer/100001.html", "spec-1", "first run"]],
    )
    write_xlsx(
        config.product_sku_master,
        ["商品SKU", "货源链接", "规格", "长", "宽", "高"],
        [["SKU000001", "https://detail.1688.com/offer/100001.html", "spec-1", "1", "2", "3"]],
    )

    first_output = process_sku_mapping(
        settings=config,
        batch_id="BATCH_FIRST",
        daily_rows=load_daily_input(config.daily_input),
        product_master_rows=load_product_sku_master(config.product_sku_master),
        uploaded_rows=[],
        historical_rows=[],
        mapping_rows=[],
    )
    first_output_dir = export_process_output(config, first_output, "BATCH_FIRST")
    publish_state_files(config, first_output_dir, "BATCH_FIRST")

    second_output = process_sku_mapping(
        settings=config,
        batch_id="BATCH_SECOND",
        daily_rows=load_daily_input(config.daily_input),
        product_master_rows=load_product_sku_master(config.product_sku_master),
        uploaded_rows=load_uploaded_product_skus(config.uploaded_product_skus),
        historical_rows=load_historical_ordered_platform_skus(config.historical_ordered_platform_skus),
        mapping_rows=load_sku_platform_mapping(config.product_sku_platform_sku_mapping),
    )

    assert second_output.summary is not None
    assert second_output.summary.first_order_processed == 0
    assert second_output.summary.historical_skipped == 1
    assert not second_output.erp_new_rows
    assert not second_output.erp_update_rows


def test_only_daily_corrected_source_is_cleaned(tmp_path: Path) -> None:
    config = settings(tmp_path)
    write_xlsx(
        config.daily_input,
        ["平台SKU", "初始商品SKU", "订单号", "平台渠道", "店铺账号", "出单时间", "校正后货源链接", "校正后规格", "备注"],
        [["PSKU_FIRST", "SKU000001", "O1", "Walmart", "shop-a", "2026-07-23 10:00:00", "https://detail.1688.com/offer/100001.html?spm=abc", "A   B----标题", ""]],
    )
    write_xlsx(
        config.product_sku_master,
        ["商品SKU", "货源链接", "规格", "长", "宽", "高"],
        [["SKU000001", "https://detail.1688.com/offer/100001.html?keep=master", "A   B----标题", "1", "2", "3"]],
    )

    daily_rows = load_daily_input(config.daily_input)
    product_rows = load_product_sku_master(config.product_sku_master)

    assert daily_rows[0].corrected_source_url == "https://detail.1688.com/offer/100001.html"
    assert daily_rows[0].corrected_spec == "A B"
    assert product_rows[0].source_url == "https://detail.1688.com/offer/100001.html?keep=master"
    assert product_rows[0].spec == "A   B----标题"


def test_daily_source_fields_are_written_to_generated_product_master(tmp_path: Path) -> None:
    config = settings(tmp_path)
    write_xlsx(
        config.daily_input,
        [
            "平台SKU",
            "初始商品SKU",
            "订单号",
            "平台渠道",
            "店铺账号",
            "出单时间",
            "校正后货源链接",
            "校正后规格",
            "一级类目",
            "类目代号",
            "图片链接",
            "采购价/￥",
            "重量/g",
            "长/cm",
            "宽/cm",
            "高/cm",
            "数量",
            "颜色",
            "材质",
            "中文报关名",
            "备注",
        ],
        [
            [
                "PSKU_GEN",
                "SKU000001",
                "O1",
                "Walmart",
                "shop-a",
                "2026-07-23 10:00:00",
                "https://detail.1688.com/offer/199999.html?spm=abc",
                "新规格----标题",
                "Home",
                "JT",
                "https://img.example.com/a.png",
                "18",
                "387",
                "71.8",
                "21.7",
                "20.6",
                "6",
                "White",
                "Plastic",
                "塑料风扇",
                "审核备注",
            ]
        ],
    )
    write_xlsx(
        config.product_sku_master,
        ["商品SKU", "货源链接", "规格", "长", "宽", "高"],
        [["SKU000001", "https://detail.1688.com/offer/100001.html", "旧规格", "1", "2", "3"]],
    )

    output = process_sku_mapping(
        settings=config,
        batch_id="20260708_000000",
        daily_rows=load_daily_input(config.daily_input),
        product_master_rows=load_product_sku_master(config.product_sku_master),
        uploaded_rows=[UploadedProductSku("SKU000001")],
        historical_rows=[],
        mapping_rows=[],
    )
    output_dir = export_process_output(config, output, "20260708_000000")
    latest_product_master = values(output_dir / "最新商品基础库留存表.xlsx")
    generated = {row[0]: row for row in latest_product_master[1:]}["JT_260708_1"]
    erp_new = values(output_dir / "ERP新增表.xlsx")
    erp_generated = {row[0]: row for row in erp_new[1:]}["JT_260708_1"]

    assert latest_product_master[0] == [
        "商品SKU",
        "货源图片链接",
        "货源链接",
        "规格",
        "采购价/￥",
        "重量/g",
        "长/cm",
        "宽/cm",
        "高/cm",
        "颜色",
        "材质",
        "数量",
        "中文报关名",
        "一级类目",
        "类目代号",
        "临时SKU",
        "供应商",
        "备注",
    ]
    assert generated[1:14] == [
        "https://img.example.com/a.png",
        "https://detail.1688.com/offer/199999.html",
        "新规格",
        "18",
        "387",
        "71.8",
        "21.7",
        "20.6",
        "White",
        "Plastic",
        "6",
        "塑料风扇",
        "Home",
    ]
    assert erp_generated[3] == "新规格"
    assert erp_generated[6] == "https://img.example.com/a.png"
    assert erp_generated[7] == "387"
    assert erp_generated[8] == "18"
    assert erp_generated[13] == "https://detail.1688.com/offer/199999.html"
    assert erp_generated[14] == "审核备注"
    assert erp_generated[16] == "塑料风扇"
    assert erp_generated[17] == "387"
    assert erp_generated[18] == "2.65"
