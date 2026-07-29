"""Validate output files generated from the full-branch SKU mapping sample."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


def rows(path: Path) -> list[list[object]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def validate(output_dir: Path) -> str:
    logs = rows(output_dir / "处理日志表.xlsx")
    exceptions = rows(output_dir / "异常待处理表.xlsx")
    erp_new = rows(output_dir / "ERP新增表.xlsx")
    erp_update = rows(output_dir / "ERP更新表.xlsx")
    history = rows(output_dir / "最新历史出单平台SKU表.xlsx")
    mapping = rows(output_dir / "最新商品SKU-平台SKU映射关系表.xlsx")
    uploaded = rows(output_dir / "最新已上传商品SKU产品表.xlsx")
    master = rows(output_dir / "最新商品基础库留存表.xlsx")

    log_by_platform = {row[2]: row for row in logs[1:]}
    expected_branches = {
        "PSKU_BR01": "初始已上传/货源无误/matched/正确SKU已上传",
        "PSKU_BR02": "初始已上传/货源有误/matched/正确SKU已上传",
        "PSKU_BR03": "初始已上传/货源有误/matched/正确SKU未上传",
        "PSKU_BR04": "初始已上传/货源有误/generated/正确SKU未上传",
        "PSKU_BR05": "初始未上传/货源无误/matched/正确SKU未上传",
        "PSKU_BR06": "初始未上传/货源有误/matched/正确SKU已上传",
        "PSKU_BR07": "初始未上传/货源有误/matched/正确SKU未上传",
        "PSKU_BR08": "初始未上传/货源有误/generated/正确SKU未上传",
        "PSKU_BR09": "初始未上传/货源有误/matched/正确SKU已上传",
        "PSKU_BR10": "初始未上传/货源有误/matched/正确SKU未上传",
        "PSKU_BR11": "初始未上传/货源有误/generated/正确SKU未上传",
        "PSKU_DEDUP": "初始未上传/货源无误/matched/正确SKU未上传",
    }
    for platform_sku, branch in expected_branches.items():
        actual = log_by_platform[platform_sku][10]
        assert actual == branch, f"{platform_sku}: {actual}"
    assert log_by_platform["PSKU_DEDUP"][5] == "O-DEDUP-EARLY"

    exception_types = {row[2]: row[10] for row in exceptions[1:]}
    assert exception_types[None] == "platform_sku_missing", exception_types
    assert exception_types["PSKU_NO_CAT"] == "missing_first_category_code", exception_types
    assert exception_types["PSKU_DUP_SOURCE"] == "ambiguous_source_key", exception_types

    new_by_sku = {row[0]: row for row in erp_new[1:]}
    update_by_sku = {row[0]: row for row in erp_update[1:]}
    assert set(update_by_sku) == {"SKU_EMPTY_MATCH_UP", "SKU_MATCH_UP", "SKU_MATCH_UP_06", "SKU_UP_CONS"}, set(update_by_sku)
    platform_to_sku = {row[1]: row[0] for row in mapping[1:]}
    generated_br04 = platform_to_sku["PSKU_BR04"]
    generated_br08 = platform_to_sku["PSKU_BR08"]
    generated_br11 = platform_to_sku["PSKU_BR11"]
    assert {
        "SKU_MATCH_NEW",
        "SKU_INIT_NEW_CONS",
        "SKU_MATCH_NEW_07",
        "SKU_EMPTY_MATCH_NEW",
        "SKU_DEDUP_EARLY",
        generated_br04,
        generated_br08,
        generated_br11,
    }.issubset(new_by_sku), set(new_by_sku)
    assert update_by_sku["SKU_UP_CONS"][1] == "OLD_UP_CONS\nPSKU_BR01"
    assert update_by_sku["SKU_EMPTY_MATCH_UP"][1] == "OLD_EMPTY_MATCH_UP\nPSKU_BR09"
    assert new_by_sku[generated_br04][1] == "PSKU_BR04"
    assert new_by_sku[generated_br08][1] == "PSKU_BR08"
    assert new_by_sku[generated_br11][1] == "PSKU_BR11"
    assert generated_br04.startswith("JT_")
    assert generated_br08.startswith("JT_")
    assert generated_br11.startswith("JT_")
    assert len({generated_br04, generated_br08, generated_br11}) == 3

    history_skus = {row[0] for row in history[1:]}
    assert "PSKU_HIST" in history_skus
    assert set(expected_branches).issubset(history_skus)
    assert "PSKU_NO_CAT" not in history_skus
    assert "PSKU_DUP_SOURCE" not in history_skus

    assert platform_to_sku["PSKU_BR02"] == "SKU_MATCH_UP"
    assert platform_to_sku["PSKU_BR04"] == generated_br04
    assert platform_to_sku["PSKU_BR08"] == generated_br08
    assert platform_to_sku["PSKU_BR11"] == generated_br11
    assert "PSKU_HIST" not in platform_to_sku
    assert "PSKU_DUP_SOURCE" not in platform_to_sku

    uploaded_skus = {row[0] for row in uploaded[1:]}
    assert {generated_br04, generated_br08, generated_br11, "SKU_EMPTY_MATCH_NEW"}.issubset(uploaded_skus)

    master_by_sku = {row[0]: row for row in master[1:]}
    assert master_by_sku["SKU_INIT_NEW_WRONG_08"][2] == "https://detail.1688.com/offer/100081.html"
    assert master_by_sku["SKU_INIT_NEW_WRONG_08"][3] == "old-08"
    assert master_by_sku[generated_br04][2] == "https://detail.1688.com/offer/100042.html"
    assert master_by_sku[generated_br08][2] == "https://detail.1688.com/offer/100082.html"
    assert master_by_sku[generated_br11][2] == "https://detail.1688.com/offer/100111.html"

    return f"full_branch_output_assertions=passed; logs={len(logs)-1}; exceptions={len(exceptions)-1}; erp_new={len(erp_new)-1}; erp_update={len(erp_update)-1}"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("output_dir", type=Path)
    args = parser.parse_args()
    print(validate(args.output_dir))


if __name__ == "__main__":
    main()
