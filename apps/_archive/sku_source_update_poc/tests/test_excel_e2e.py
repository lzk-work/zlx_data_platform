from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from apps.sku_mapping_poc.src.models import ProductSkuMaster
from apps.sku_source_update_poc.src.exporter import export_output
from apps.sku_source_update_poc.src.loader import load_input_rows
from apps.sku_source_update_poc.src.models import AppSettings
from apps.sku_source_update_poc.src.processor import process_source_update


def write_xlsx(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def values(path: Path) -> list[list[str | None]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def test_excel_input_exports_mapping_and_new_product(tmp_path: Path) -> None:
    input_path = tmp_path / "input.xlsx"
    write_xlsx(
        input_path,
        ["平台SKU", "初始商品SKU", "校正后货源链接", "校正后规格", "类目代号", "采购价/￥", "重量/g", "备注"],
        [
            ["PSKU_MATCH", "INIT1", "https://detail.1688.com/offer/100001.html?spm=abc", "spec-1", "", "10", "0.1", "match"],
            ["PSKU_NEW", "INIT2", "https://detail.1688.com/offer/199999.html?spm=abc", "spec-new----标题", "JT", "12", "0.2", "new"],
        ],
    )
    config = AppSettings(tmp_path, input_path, tmp_path / "output", "postgresql://user:pass@localhost:5432/db")
    output = process_source_update(
        settings=config,
        batch_id="20260725_120000",
        input_rows=load_input_rows(input_path),
        product_source_rows=[ProductSkuMaster("SKU_EXIST", "https://detail.1688.com/offer/100001.html", "spec-1")],
        first_category_code_rows=[],
    )

    output_dir = export_output(config, output, "20260725_120000")

    mapping_rows = values(output_dir / "本次平台SKU映射关系表.xlsx")
    assert mapping_rows[1][0:6] == ["PSKU_MATCH", "INIT1", "SKU_EXIST", "https://detail.1688.com/offer/100001.html", "spec-1", "matched_existing"]
    assert mapping_rows[2][0:6] == ["PSKU_NEW", "INIT2", "JT_260725_1", "https://detail.1688.com/offer/199999.html", "spec-new", "generated_new"]

    new_products = values(output_dir / "本次商品基础库新增表.xlsx")
    assert new_products[1][0] == "JT_260725_1"
    assert new_products[1][2] == "https://detail.1688.com/offer/199999.html"
    assert new_products[1][3] == "spec-new"
    assert new_products[1][5] == "0.2"


def test_source_only_excel_input_exports_new_products_without_mapping(tmp_path: Path) -> None:
    input_path = tmp_path / "input.xlsx"
    write_xlsx(
        input_path,
        ["平台SKU", "初始商品SKU", "校正后货源链接", "校正后规格", "类目代号", "备注"],
        [
            ["", "", "https://detail.1688.com/offer/100001.html", "spec-1", "", "existing"],
            ["", "", "https://detail.1688.com/offer/199999.html", "spec-new", "JT", "new"],
        ],
    )
    config = AppSettings(tmp_path, input_path, tmp_path / "output", "postgresql://user:pass@localhost:5432/db", run_mode="source-only")
    output = process_source_update(
        settings=config,
        batch_id="20260725_120000",
        input_rows=load_input_rows(input_path),
        product_source_rows=[ProductSkuMaster("SKU_EXIST", "https://detail.1688.com/offer/100001.html", "spec-1")],
        first_category_code_rows=[],
    )

    output_dir = export_output(config, output, "20260725_120000")

    assert output.summary is not None
    assert output.summary.source_only_skipped == 1
    assert output.summary.generated_product_skus == 1
    assert values(output_dir / "本次平台SKU映射关系表.xlsx") == [["平台SKU", "初始商品SKU", "正确商品SKU", "校正后货源链接", "校正后规格", "匹配结果", "备注"]]
    assert values(output_dir / "本次商品基础库新增表.xlsx")[1][0] == "JT_260725_1"
