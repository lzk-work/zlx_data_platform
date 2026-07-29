"""平台 SKU 货源预校正 Excel 读取。"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from apps.sku_mapping_poc.src.normalizer import clean_source_url, clean_spec, normalize_sku, normalize_text

from .models import SourceUpdateInputRow, WorkbookRow

INPUT_COLUMNS = ["平台SKU", "初始商品SKU", "校正后货源链接", "校正后规格"]


def load_input_rows(path: Path, sheet_name: str | None = None) -> list[SourceUpdateInputRow]:
    return [_input_row(row_number, row) for row_number, row in _read_rows(path, INPUT_COLUMNS, sheet_name)]


def _read_rows(path: Path, required_columns: list[str], sheet_name: str | None = None) -> list[tuple[int, WorkbookRow]]:
    if not path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {path}")
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Excel 文件没有表头: {path}")

    headers = [normalize_text(value) for value in rows[0]]
    missing = [column for column in required_columns if column not in headers]
    if missing:
        raise ValueError(f"{path} 缺少字段: {', '.join(missing)}")

    result: list[tuple[int, WorkbookRow]] = []
    for index, values in enumerate(rows[1:], start=2):
        row = {header: value for header, value in zip(headers, values, strict=False) if header}
        if any(normalize_text(row.get(column)) for column in required_columns):
            result.append((index, row))
    return result


def _input_row(row_number: int, row: WorkbookRow) -> SourceUpdateInputRow:
    return SourceUpdateInputRow(
        row_number=row_number,
        platform_sku=normalize_sku(row.get("平台SKU")),
        initial_product_sku=normalize_sku(row.get("初始商品SKU")),
        corrected_source_url=clean_source_url(row.get("校正后货源链接")),
        corrected_spec=clean_spec(row.get("校正后规格")),
        first_level_category=normalize_text(row.get("一级类目")),
        category_code=normalize_text(row.get("类目代号")),
        source_image_url=normalize_text(row.get("图片链接") or row.get("货源图片链接")),
        purchase_price=normalize_text(row.get("采购价/￥")),
        weight_g=normalize_text(row.get("重量/g")),
        length_cm=normalize_text(row.get("长/cm") or row.get("长")),
        width_cm=normalize_text(row.get("宽/cm") or row.get("宽")),
        height_cm=normalize_text(row.get("高/cm") or row.get("高")),
        color=normalize_text(row.get("颜色")),
        material=normalize_text(row.get("材质")),
        quantity=normalize_text(row.get("数量")),
        chinese_customs_name=normalize_text(row.get("中文报关名")),
        supplier=normalize_text(row.get("供应商")),
        remark=normalize_text(row.get("备注")),
    )
