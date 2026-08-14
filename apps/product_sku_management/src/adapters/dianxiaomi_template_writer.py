"""Writers that preserve Dianxiaomi workbook headers and unused columns."""

from __future__ import annotations

from copy import copy
from pathlib import Path
import re
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook


def write_template_rows(template_path: str | Path, output_path: str | Path, rows: Iterable[dict[str, Any]]) -> None:
    """按店小秘模板表头写入数据。

    Args:
        template_path: 店小秘原始模板路径。
        output_path: 写入后的输出文件路径。
        rows: 待写入行；只填充能匹配到模板表头的字段。

    Returns:
        None: 文件写入到output_path，未匹配列保持空值和原表头格式。
    """
    template_workbook = load_workbook(template_path)
    template_sheet = template_workbook.worksheets[0]
    headers = [str(cell.value or "").strip() for cell in template_sheet[1]]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = template_sheet.title
    copy_header_row(template_sheet, sheet)

    for row in rows:
        sheet.append([row.get(header, row.get(normalize_header_key(header), "")) for header in headers])

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    template_workbook.close()
    workbook.close()


def copy_header_row(source_sheet: Any, target_sheet: Any) -> None:
    """复制店小秘模板首行表头和基础表头格式。

    Args:
        source_sheet: 原始模板sheet。
        target_sheet: 新输出sheet。

    Returns:
        None: 目标sheet首行拥有与模板一致的表头值和样式。
    """
    for cell in source_sheet[1]:
        target = target_sheet.cell(row=1, column=cell.column, value=cell.value)
        if cell.has_style:
            target.font = copy(cell.font)
            target.fill = copy(cell.fill)
            target.border = copy(cell.border)
            target.alignment = copy(cell.alignment)
            target.number_format = cell.number_format
            target.protection = copy(cell.protection)
        if cell.comment:
            target.comment = copy(cell.comment)

    if source_sheet.row_dimensions[1].height is not None:
        target_sheet.row_dimensions[1].height = source_sheet.row_dimensions[1].height
    for key, dimension in source_sheet.column_dimensions.items():
        if dimension.width is not None:
            target_sheet.column_dimensions[key].width = dimension.width
    target_sheet.freeze_panes = source_sheet.freeze_panes
    if source_sheet.auto_filter.ref:
        target_sheet.auto_filter.ref = source_sheet.auto_filter.ref


def normalize_header_key(header: str) -> str:
    """标准化模板表头键。

    Args:
        header: 模板中的原始表头。

    Returns:
        str: 移除空白和换行后的表头键。
    """
    return re.sub(r"\s+", "", header)
