"""Excel readers for product SKU management."""

from __future__ import annotations

import re
import unicodedata
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..models.input_models import PlatformListingInputRow, SourceGroupInput

SOURCE_LINK_HEADER_PATTERN = re.compile(r"^货源链接(\d+)$")


def read_platform_listing_rows(path: str | Path) -> list[PlatformListingInputRow]:
    """读取平台SKU补充输入表。

    Args:
        path: Excel输入文件路径，读取第一个sheet。

    Returns:
        list[PlatformListingInputRow]: 已按表头转换后的输入行列表。
    """
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.worksheets[0]
    headers = [normalize_header(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    header_index = {header: index for index, header in enumerate(headers) if header}
    rows: list[PlatformListingInputRow] = []

    for row_no, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if is_empty_row(row):
            continue
        raw_row = {header: row[index] for header, index in header_index.items() if index < len(row)}
        rows.append(
            PlatformListingInputRow(
                row_no=row_no,
                shop_name=cell_text(raw_row.get("店铺")),
                platform_sku=cell_text(raw_row.get("平台SKU")),
                first_level_category=cell_text(raw_row.get("一级类目")),
                main_image_url=cell_text(raw_row.get("主图链接")),
                length_cm=optional_dimension(raw_row.get("长/cm"), "长/cm"),
                width_cm=optional_dimension(raw_row.get("宽/cm"), "宽/cm"),
                height_cm=optional_dimension(raw_row.get("高/cm"), "高/cm"),
                logistics_attribute=cell_text(raw_row.get("属性")),
                chinese_customs_name=cell_text(raw_row.get("中文报关名")),
                development_note=cell_text(raw_row.get("开发备注")),
                source_groups=tuple(read_source_groups(raw_row)),
                raw_row=raw_row,
            )
        )

    workbook.close()
    return rows


def read_source_groups(raw_row: dict[str, Any]) -> list[SourceGroupInput]:
    """读取一行里的货源组。

    Args:
        raw_row: 按表头映射后的原始Excel行。

    Returns:
        list[SourceGroupInput]: 非空的货源组列表，空组会跳过。
    """
    group_numbers = sorted(
        int(match.group(1))
        for key in raw_row
        if (match := SOURCE_LINK_HEADER_PATTERN.fullmatch(key))
    )
    groups: list[SourceGroupInput] = []
    for group_no in group_numbers:
        source_url = cell_text(raw_row.get(f"货源链接{group_no}"))
        spec = cell_text(raw_row.get(f"货源{group_no}规格"))
        note = cell_text(raw_row.get(f"货源{group_no}备注"))
        price = raw_row.get(f"货源{group_no}采购价")
        weight = raw_row.get(f"货源{group_no}重量/kg")
        if not any([source_url, spec, note, price not in (None, ""), weight not in (None, "")]):
            continue
        groups.append(
            SourceGroupInput(
                group_no=group_no,
                source_url=source_url,
                spec=spec,
                note=note,
                purchase_price_rmb=optional_decimal(price, f"货源{group_no}采购价"),
                weight_kg=optional_decimal(weight, f"货源{group_no}重量/kg"),
            )
        )
    return groups


def normalize_header(value: object) -> str:
    """标准化Excel表头文本。

    Args:
        value: 单元格原始值。

    Returns:
        str: 去除首尾空白后的表头文本。
    """
    return str(value or "").strip()


def cell_text(value: object) -> str:
    """将单元格值转换为文本。

    Args:
        value: 单元格原始值。

    Returns:
        str: 去除首尾空白后的文本；空值返回空字符串。
    """
    if value is None:
        return ""
    return str(value).strip()


def optional_decimal(value: object, field_name: str) -> Decimal | None:
    """转换可为空的数字单元格。

    Args:
        value: 单元格原始值。
        field_name: 字段名称，用于异常提示。

    Returns:
        Decimal | None: 数字值；空单元格返回None。

    Raises:
        ValueError: 非数字内容无法转换时抛出。
    """
    if value in (None, ""):
        return None
    decimal_text = normalize_decimal_text(value)
    if not decimal_text:
        return None
    try:
        return Decimal(decimal_text)
    except (InvalidOperation, AttributeError) as exc:
        raise ValueError(f"{field_name}必须是数字") from exc


def normalize_decimal_text(value: object) -> str:
    """清理Excel数字文本中的不可见格式字符。

    Args:
        value: 单元格原始值。

    Returns:
        str: 去除首尾空白和Unicode格式控制字符后的数字文本。
    """
    text = str(value).strip()
    return "".join(char for char in text if unicodedata.category(char) != "Cf").strip()


def optional_dimension(value: object, field_name: str) -> Decimal | None:
    """转换可为空的尺寸单元格。

    Args:
        value: 尺寸单元格原始值。
        field_name: 字段名称，用于异常提示。

    Returns:
        Decimal | None: 大于0的尺寸值；空值或0返回None。

    Raises:
        ValueError: 非数字内容无法转换时抛出。
    """
    decimal_value = optional_decimal(value, field_name)
    if decimal_value is None:
        return None
    if decimal_value == Decimal("0"):
        return None
    return decimal_value.quantize(Decimal("0.0001"))


def is_empty_row(row: tuple[Any, ...]) -> bool:
    """判断Excel行是否为空行。

    Args:
        row: Excel一行的单元格值。

    Returns:
        bool: 所有单元格都为空时返回True。
    """
    return all(value is None or str(value).strip() == "" for value in row)
