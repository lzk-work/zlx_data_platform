"""Dianxiaomi bundle SKU exporter tests."""

from decimal import Decimal

from openpyxl import Workbook, load_workbook

from apps.product_sku_management.src.exporters.dianxiaomi_bundle_sku_exporter import export_bundle_sku_template
from apps.product_sku_management.src.models.domain_models import BundleSkuRecord


def test_export_bundle_sku_template_writes_sales_unit_dimensions(tmp_path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "bundle.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "*组合sku",
            "包含的商品sku",
            "数量",
            "申报重量(g)",
            "长（cm）",
            "宽（cm）",
            "高（cm）",
            "申报金额（USD）",
            "危险运输品",
        ]
    )
    workbook.save(template_path)

    export_bundle_sku_template(
        template_path,
        output_path,
        [
            BundleSkuRecord(
                bundle_sku="ZH_260731_1_2_1",
                bundle_name="2 个/组，红色*2",
                total_product_count=2,
                distinct_product_sku_count=1,
                items=(("YS_260731_1", 2),),
                main_image_url="https://example.com/a.jpg",
                chinese_customs_name="玩具",
                reference_total_purchase_price_rmb=Decimal("20"),
                reference_total_weight_g=Decimal("300.235"),
                logistics_attribute="带电",
                note="",
                length_cm=Decimal("10.235"),
                width_cm=Decimal("8.234"),
                height_cm=Decimal("3.235"),
            )
        ],
        exchange_rate_usd=6.8,
    )

    saved = load_workbook(output_path, data_only=True)
    row = [cell.value for cell in saved.active[2]]
    saved.close()

    assert row[:3] == ["ZH_260731_1_2_1", "YS_260731_1", 2]
    assert Decimal(str(row[3])) == Decimal("300.24")
    assert Decimal(str(row[4])) == Decimal("10.24")
    assert Decimal(str(row[5])) == Decimal("8.23")
    assert Decimal(str(row[6])) == Decimal("3.24")
    assert Decimal(str(row[7])) == Decimal("2.94")
    assert row[8] == "1"


def test_export_bundle_sku_template_applies_min_declared_amount(tmp_path) -> None:
    """组合SKU整组采购价过低时，申报金额（USD）按店小秘下限0.1导出。"""
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "bundle.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["*组合sku", "包含的商品sku", "数量", "申报金额（USD）"])
    workbook.save(template_path)

    export_bundle_sku_template(
        template_path,
        output_path,
        [
            BundleSkuRecord(
                bundle_sku="ZH_260814_1_2_1",
                bundle_name="2 个/组，红色*2",
                total_product_count=2,
                distinct_product_sku_count=1,
                items=(("YS_260814_1", 2),),
                main_image_url="https://example.com/a.jpg",
                chinese_customs_name="玩具",
                reference_total_purchase_price_rmb=Decimal("0.4"),
                reference_total_weight_g=Decimal("100"),
                logistics_attribute="普货",
                note="",
            )
        ],
        exchange_rate_usd=6.8,
    )

    saved = load_workbook(output_path, data_only=True)
    row = [cell.value for cell in saved.active[2]]
    saved.close()

    # 0.4 / 6.8 ≈ 0.06 低于下限，申报金额按 0.1 输出
    assert Decimal(str(row[3])) == Decimal("0.1")


def test_export_bundle_sku_template_repeats_bundle_sku_for_each_item_row(tmp_path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "bundle.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["*组合sku", "包含的商品sku", "数量"])
    workbook.save(template_path)

    export_bundle_sku_template(
        template_path,
        output_path,
        [
            BundleSkuRecord(
                bundle_sku="ZH_260731_2_2_1",
                bundle_name="2 个/组，红色---数量1，蓝色---数量1",
                total_product_count=2,
                distinct_product_sku_count=2,
                items=(("YS_260731_1", 1), ("YS_260731_2", 1)),
                main_image_url="https://example.com/a.jpg",
                chinese_customs_name="玩具",
                reference_total_purchase_price_rmb=Decimal("20"),
                reference_total_weight_g=Decimal("300"),
                logistics_attribute="普货",
                note="",
            )
        ],
        exchange_rate_usd=7,
    )

    saved = load_workbook(output_path, data_only=True)
    rows = [[cell.value for cell in row] for row in saved.active.iter_rows(min_row=2, max_row=3)]
    saved.close()

    assert rows == [
        ["ZH_260731_2_2_1", "YS_260731_1", 1],
        ["ZH_260731_2_2_1", "YS_260731_2", 1],
    ]


def test_export_bundle_sku_template_writes_deduped_source_urls(tmp_path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "bundle.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["*组合sku", "来源URL(必须以http://或https://开头)", "包含的商品sku", "数量"])
    workbook.save(template_path)

    export_bundle_sku_template(
        template_path,
        output_path,
        [
            BundleSkuRecord(
                bundle_sku="ZH_260731_2_2_1",
                bundle_name="2 个/组，红色---数量1，蓝色---数量1",
                total_product_count=2,
                distinct_product_sku_count=2,
                items=(("YS_260731_1", 1), ("YS_260731_2", 1)),
                main_image_url="https://example.com/a.jpg",
                chinese_customs_name="玩具",
                reference_total_purchase_price_rmb=Decimal("20"),
                reference_total_weight_g=Decimal("300"),
                logistics_attribute="普货",
                note="",
                source_urls=(
                    "https://detail.1688.com/offer/111.html",
                    "https://detail.1688.com/offer/222.html",
                    "https://detail.1688.com/offer/111.html",
                ),
            )
        ],
        exchange_rate_usd=7,
    )

    saved = load_workbook(output_path, data_only=True)
    rows = [[cell.value for cell in row] for row in saved.active.iter_rows(min_row=2, max_row=3)]
    saved.close()

    assert rows == [
        [
            "ZH_260731_2_2_1",
            "https://detail.1688.com/offer/111.html\nhttps://detail.1688.com/offer/222.html",
            "YS_260731_1",
            1,
        ],
        ["ZH_260731_2_2_1", None, "YS_260731_2", 1],
    ]
