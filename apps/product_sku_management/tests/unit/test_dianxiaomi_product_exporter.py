"""Dianxiaomi product SKU exporter tests."""

from decimal import Decimal

from openpyxl import Workbook, load_workbook

from apps.product_sku_management.src.exporters.dianxiaomi_product_sku_exporter import export_product_sku_template
from apps.product_sku_management.src.models.domain_models import ProductSkuRecord


def test_export_product_sku_template_writes_dangerous_transport_code(tmp_path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "product.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "*SKU(必填)",
            "中文名称",
            "商品净重（g）",
            "长（cm）",
            "宽（cm）",
            "高（cm）",
            "申报重量(g)",
            "申报金额（USD）",
            "危险运输品",
        ]
    )
    workbook.save(template_path)

    export_product_sku_template(
        template_path,
        output_path,
        [
            ProductSkuRecord(
                product_sku="YS_260731_1",
                source_url="https://detail.1688.com/offer/1.html",
                source_platform="1688",
                spec="白色",
                quantity=1,
                product_sku_type="normal",
                package_fingerprint=None,
                package_details=(),
                product_name="白色---数量1",
                main_image_url="https://example.com/a.jpg",
                first_level_category="Fashion",
                category_code="YS",
                reference_purchase_price_rmb=Decimal("10"),
                reference_weight_g=Decimal("100.235"),
                chinese_customs_name="饰品",
                logistics_attribute="普货",
                note="",
                length_cm=Decimal("10.235"),
                width_cm=Decimal("8.234"),
                height_cm=Decimal("3.235"),
                is_direct_sales_unit=True,
            )
        ],
        exchange_rate_usd=6.8,
    )

    saved = load_workbook(output_path, data_only=True)
    row = [cell.value for cell in saved.active[2]]
    saved.close()

    assert row[:2] == ["YS_260731_1", "白色---数量1"]
    assert Decimal(str(row[2])) == Decimal("100.24")
    assert Decimal(str(row[3])) == Decimal("10.24")
    assert Decimal(str(row[4])) == Decimal("8.23")
    assert Decimal(str(row[5])) == Decimal("3.24")
    assert Decimal(str(row[6])) == Decimal("100.24")
    assert Decimal(str(row[7])) == Decimal("1.47")
    assert row[8] == "0"


def test_export_product_sku_template_writes_forced_package_source_urls(tmp_path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "product.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["*SKU(必填)", "来源URL（必须以http://或https：//开头）", "备注"])
    workbook.save(template_path)

    export_product_sku_template(
        template_path,
        output_path,
        [
            ProductSkuRecord(
                product_sku="YS_260731_2",
                source_url="https://detail.1688.com/offer/1.html",
                source_platform="1688",
                spec="红色",
                quantity=1,
                product_sku_type="forced_package",
                package_fingerprint="fp",
                package_details=(
                    {"source_url": "https://detail.1688.com/offer/1.html"},
                    {"source_url": "https://detail.1688.com/offer/2.html"},
                    {"source_url": "https://detail.1688.com/offer/1.html"},
                ),
                product_name="2个/组，红色---数量1，蓝色---数量1",
                main_image_url="https://example.com/a.jpg",
                first_level_category="Fashion",
                category_code="YS",
                reference_purchase_price_rmb=Decimal("10"),
                reference_weight_g=Decimal("100"),
                chinese_customs_name="饰品",
                logistics_attribute="普货",
                note="原备注\n强制合并",
                is_direct_sales_unit=True,
            )
        ],
        exchange_rate_usd=7,
    )

    saved = load_workbook(output_path, data_only=True)
    row = [cell.value for cell in saved.active[2]]
    saved.close()

    assert row == [
        "YS_260731_2",
        "https://detail.1688.com/offer/1.html\nhttps://detail.1688.com/offer/2.html",
        "原备注\n强制合并",
    ]


def test_export_product_sku_template_uses_package_weight_and_price_for_multi_piece(tmp_path) -> None:
    """多件单品(3件/包)导出店小秘时，重量与采购价应按整包(单品值×数量)输出。"""
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "product.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "*SKU(必填)",
            "商品净重（g）",
            "采购参考价（RMB）",
            "申报重量(g)",
            "申报金额（USD）",
        ]
    )
    workbook.save(template_path)

    export_product_sku_template(
        template_path,
        output_path,
        [
            ProductSkuRecord(
                product_sku="YS_260814_3",
                source_url="https://detail.1688.com/offer/3.html",
                source_platform="1688",
                spec="蓝色",
                quantity=3,
                product_sku_type="normal",
                package_fingerprint=None,
                package_details=(),
                product_name="蓝色---数量3",
                main_image_url="https://example.com/c.jpg",
                first_level_category="Fashion",
                category_code="YS",
                reference_purchase_price_rmb=Decimal("10"),
                reference_weight_g=Decimal("500.235"),
                chinese_customs_name="饰品",
                logistics_attribute="普货",
                note="",
                is_direct_sales_unit=True,
            )
        ],
        exchange_rate_usd=7,
    )

    saved = load_workbook(output_path, data_only=True)
    row = [cell.value for cell in saved.active[2]]
    saved.close()

    # 整包 = 单品值 × 数量：重量 500.235×3≈1500.71→1500.71，采购价 10×3=30，申报金额 30/7≈4.29
    assert Decimal(str(row[1])) == Decimal("1500.71")
    assert Decimal(str(row[2])) == Decimal("30")
    assert Decimal(str(row[3])) == Decimal("1500.71")
    assert Decimal(str(row[4])) == Decimal("4.29")


def test_export_product_sku_template_applies_min_declared_amount(tmp_path) -> None:
    """整包采购价过低时，申报金额（USD）按店小秘下限0.1导出，采购参考价保持原值。"""
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "product.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["*SKU(必填)", "采购参考价（RMB）", "申报金额（USD）"])
    workbook.save(template_path)

    export_product_sku_template(
        template_path,
        output_path,
        [
            ProductSkuRecord(
                product_sku="YS_260814_4",
                source_url="https://detail.1688.com/offer/4.html",
                source_platform="1688",
                spec="白色",
                quantity=1,
                product_sku_type="normal",
                package_fingerprint=None,
                package_details=(),
                product_name="白色---数量1",
                main_image_url="https://example.com/d.jpg",
                first_level_category="Fashion",
                category_code="YS",
                reference_purchase_price_rmb=Decimal("0.2"),
                reference_weight_g=Decimal("50"),
                chinese_customs_name="饰品",
                logistics_attribute="普货",
                note="",
                is_direct_sales_unit=True,
            )
        ],
        exchange_rate_usd=6.8,
    )

    saved = load_workbook(output_path, data_only=True)
    row = [cell.value for cell in saved.active[2]]
    saved.close()

    # 0.2 / 6.8 ≈ 0.03 低于下限，申报金额按 0.1 输出；采购参考价仍为整包原值 0.2
    assert Decimal(str(row[1])) == Decimal("0.2")
    assert Decimal(str(row[2])) == Decimal("0.1")
